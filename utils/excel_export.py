import io
import re

import pandas as pd
from openpyxl.styles import PatternFill


_CARACTERES_PROHIBIDOS_HOJA = re.compile(r'[:\\/?*\[\]]')

_COLOR_RESALTADO = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")


def nombre_hoja_valido(nombre, nombres_usados):
    """Convierte un nombre de columna en un nombre de hoja Excel válido:
    sin los caracteres prohibidos por Excel (: \\ / ? * [ ]), máximo 31
    caracteres, y sin colisionar con nombres ya usados en el mismo libro."""

    nombre_limpio = _CARACTERES_PROHIBIDOS_HOJA.sub("_", str(nombre))
    nombre_limpio = nombre_limpio[:31].strip() or "hoja"

    nombre_final = nombre_limpio
    contador = 1

    while nombre_final in nombres_usados:
        sufijo = f"_{contador}"
        nombre_final = nombre_limpio[:31 - len(sufijo)] + sufijo
        contador += 1

    nombres_usados.add(nombre_final)
    return nombre_final


def generar_excel_consolidado(resultados_por_columna):
    """Recibe un diccionario {nombre_columna: DataFrame} y devuelve los
    bytes de un Excel con una hoja por cada columna. En cada hoja, la
    columna que originó el análisis queda resaltada en amarillo para
    ubicarla rápidamente entre el resto de la fila."""

    buffer = io.BytesIO()
    nombres_usados = set()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        for nombre_columna, df_resultado in resultados_por_columna.items():

            nombre_hoja = nombre_hoja_valido(nombre_columna, nombres_usados)

            # Quitar fila_excel antes de exportar — es una columna interna
            # de referencia, no parte del archivo original del usuario
            df_exportar = df_resultado.drop(columns=["fila_excel"], errors="ignore")

            df_exportar.to_excel(writer, sheet_name=nombre_hoja, index=False)

            worksheet = writer.sheets[nombre_hoja]

            if nombre_columna in df_exportar.columns:

                indice_columna = df_exportar.columns.get_loc(nombre_columna) + 1
                total_filas = len(df_exportar)

                for fila in range(1, total_filas + 2):  # +2: encabezado + 1-indexed
                    worksheet.cell(row=fila, column=indice_columna).fill = _COLOR_RESALTADO

    return buffer.getvalue()